import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Get today's date and calculate Monday and Saturday of this week
today = datetime.today()
monday = today - timedelta(days=today.weekday())
saturday = monday + timedelta(days=5)
month_str = monday.strftime('%B').upper()

# Paths
downloads_path = "C:/Downloads"
report_dir = "C:/Users/MAY"
report_filename = f"TM MY Income Per Hour Report {monday.day}th - {saturday.day}th {month_str} 2025.xlsx"
report_path = os.path.join(report_dir, report_filename)

# Load target workbook and worksheet
target_wb = load_workbook(report_path)
target_ws = target_wb["Working Days Raw"]

# ✅ Unmerge all merged cells before deleting rows
if target_ws.merged_cells.ranges:
    print(f"Unmerging {len(target_ws.merged_cells.ranges)} merged cell ranges...")
    merged_ranges = list(target_ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        target_ws.unmerge_cells(str(merged_range))

# ✅ Delete all rows
target_ws.delete_rows(1, target_ws.max_row)

# Look for matching files in Downloads
for file in os.listdir(downloads_path):
    if file.endswith(".xlsx") and not file.startswith("~$"):
        file_path = os.path.join(downloads_path, file)
        try:
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws["A3"].value == "Agent System Activities" and ws["E3"].value == "Kavitha Malaysia":
                    print(f"✅ Copying from: {file}, Sheet: {sheet_name}")
                    for row in ws.iter_rows(values_only=True):
                        target_ws.append(row)
        except Exception as e:
            print(f"❌ Failed to process {file}: {e}")

# Save updated workbook
target_wb.save(report_path)
print(f"✅ Data replaced successfully in: {report_path}")
