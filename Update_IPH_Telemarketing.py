import openpyxl
import os
import datetime
from openpyxl import Workbook
from copy import copy

def get_day_sheet_from_date(date_value, file_type=None):
    if file_type == "activities":
        return "Working Days Raw"
    days_map = {
        "Monday": "MON RAW",
        "Tuesday": "TUE RAW",
        "Wednesday": "WED RAW",
        "Thursday": "THU RAW",
        "Friday": "FRI RAW",
        "Saturday": "SAT RAW"  
    }
    return days_map.get(date_value.strftime('%A'), None)

def check_file_content(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        print(f"\U0001F50E Checking {file_path}: A3='{ws['A3'].value}', E3='{ws['E3'].value}', A5='{ws['A5'].value}', B3='{ws['B3'].value}'")

        date_value = ws["B3"].value
        if isinstance(date_value, datetime.datetime):
            date_value = date_value.date()
        elif isinstance(date_value, str):
            try:
                date_value = datetime.datetime.strptime(date_value, "%Y-%m-%d").date()
            except ValueError:
                print(f"❗ Error parsing date from string: {date_value}")
                return None, None

        if (
            ws["A3"].value == "Agent Calling Activity - Consolidated" and
            ws["E3"].value == "Agent Name" and
            ws["A5"].value == "Initiative Name"
        ):
            wb.close()
            return date_value, "som"

        elif (
            ws["A3"].value == "Agent Calling Activity - Consolidated" and
            ws["E3"].value == "Agent Name" and
            ws["A5"].value == "Initiative Name"
        ):
            wb.close()
            return date_value, "initiative"

        elif (
            ws["A3"].value == "Agent Calling Activity - Consolidated" and
            ws["E3"].value == "Agent Name" and
            ws["A5"].value == "Agent Name"
        ):
            wb.close()
            return date_value, "individual"

        elif (
            ws["A3"].value == "Agent System Activities" and
            ws["E3"].value == "Manager Name"
        ):
            wb.close()
            return date_value, "activities"

        wb.close()
    except Exception as e:
        print(f"❗ Error reading {file_path}: {e}")
    return None, None

def copy_data(source_file, report_wb, sheet_name):
    try:
        src_wb = openpyxl.load_workbook(source_file)
        src_ws = src_wb.active

        if sheet_name not in report_wb.sheetnames:
            print(f"❌ Error: Sheet '{sheet_name}' not found in the report file.")
            return

        dst_ws = report_wb[sheet_name]
        print(f"📥 Copying data to sheet: {sheet_name}")

        # Unmerge cells in destination before pasting data
        for merged_range in list(dst_ws.merged_cells.ranges):
            dst_ws.unmerge_cells(str(merged_range))

        # Paste data with styles
        for row_idx, row in enumerate(src_ws.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                try:
                    new_cell = dst_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.alignment = copy(cell.alignment)
                except Exception as cell_error:
                    print(f"❌ Error at cell ({row_idx}, {col_idx}) with value '{cell.value}': {cell_error}")
                    raise

        src_wb.close()
    except Exception as e:
        print(f"❌ Error copying data from {source_file} to sheet '{sheet_name}': {e}")

def find_all_valid_files(download_folder):
    matched_files = []
    for file in os.listdir(download_folder):
        if file.endswith(".xlsx") and not file.startswith("~$"):
            file_path = os.path.join(download_folder, file)
            date_value, file_type = check_file_content(file_path)
            if date_value:
                matched_files.append((file_path, date_value, file_type))
                print(f"✅ Matching file found: {file_path} with date {date_value.strftime('%A, %d/%m/%Y')} - Type: {file_type}")
    return matched_files

def get_weekly_report_path(file_type, date_value):
    monday = date_value - datetime.timedelta(days=date_value.weekday())
    saturday = monday + datetime.timedelta(days=5)
    month_str = date_value.strftime('%B').upper()  # replaces hardcoded "MAY"

    if file_type == "Agent_1":
        filename = f"SOM Income Per Hour Report {monday.day}th - {saturday.day}th {month_str} 2025 (by Initiative).xlsx"
        return os.path.join("C:/Users/TELEMARKETING/Initiative/2025/Agent_1/MAY", filename)

    folder_base = "C:/Users/TELEMARKETING"
    if file_type in ["individual", "activities"]:
        subfolder = "Individual"
        filename = f"TM MY Income Per Hour Report {monday.day}th - {saturday.day}th {month_str} 2025.xlsx"
    elif file_type == "initiative":
        subfolder = "Initiative"
        filename = f"TM MY Income Per Hour Report {monday.day}th - {saturday.day}th {month_str} 2025 (by initiative).xlsx"
    else:
        return None

    folder_path = os.path.join(folder_base, subfolder, f"2025/{month_str}")
    return os.path.join(folder_path, filename)

def main():
    download_folder = "C:/Users/Downloads"
    matched_files = find_all_valid_files(download_folder)
    if not matched_files:
        print("⚠️ No matching files found in Downloads folder. No update performed.")
        return

    workbooks = {}

    for file_path, date_value, file_type in matched_files:
        report_path = get_weekly_report_path(file_type, date_value)
        if not report_path:
            print(f"❌ Unknown file type: {file_type}. Skipping {file_path}")
            continue

        if report_path not in workbooks:
            if os.path.exists(report_path):
                try:
                    workbooks[report_path] = openpyxl.load_workbook(report_path)
                except Exception as e:
                    print(f"❌ Error loading report file at {report_path}: {e}")
                    continue
            else:
                print(f"❌ Report not found at: {report_path}. Skipping update.")
                continue

        sheet_name = get_day_sheet_from_date(date_value, file_type)
        if sheet_name:
            print(f"📝 Updating {sheet_name} in report with data from {file_path}")
            copy_data(file_path, workbooks[report_path], sheet_name)
        else:
            print(f"❌ Skipping {file_path}: could not determine appropriate sheet for {date_value}")

    for report_path, wb in workbooks.items():
        wb.save(report_path)
        wb.close()
        print(f"✅ Report saved successfully at: {report_path}")

if __name__ == "__main__":
    main()
